import csv
import io
import json
import uuid
import re
import requests
from decimal import Decimal
from typing import List, Tuple, Dict, Any, Optional
from openpyxl import load_workbook
from django.conf import settings
from qdrant_client import QdrantClient
from qdrant_client.models import Distance, VectorParams, PointStruct

from apps.ai_customer.models import AICustomerSetting


def parse_text_file(file_name: str, raw: bytes) -> str:
    lower = file_name.lower()
    if lower.endswith(".json"):
        data = json.loads(raw.decode("utf-8"))
        if isinstance(data, list):
            return "\n".join(json.dumps(item, ensure_ascii=False) for item in data)
        if isinstance(data, dict):
            return json.dumps(data, ensure_ascii=False)
        return str(data)

    if lower.endswith(".jsonl"):
        lines = []
        text = raw.decode("utf-8", errors="ignore")
        for idx, row in enumerate(text.splitlines(), start=1):
            row = row.strip()
            if not row:
                continue
            try:
                obj = json.loads(row)
            except json.JSONDecodeError as exc:
                raise ValueError(f"jsonl第{idx}行格式错误: {exc}")
            if isinstance(obj, (dict, list)):
                lines.append(json.dumps(obj, ensure_ascii=False))
            else:
                lines.append(str(obj))
        return "\n".join(lines)

    if lower.endswith(".csv"):
        rows = []
        with io.StringIO(raw.decode("utf-8", errors="ignore")) as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(json.dumps(row, ensure_ascii=False))
        return "\n".join(rows)

    if lower.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        lines = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                values = [str(v).strip() for v in row if v is not None and str(v).strip()]
                if values:
                    lines.append(" | ".join(values))
        return "\n".join(lines)

    if lower.endswith(".txt") or lower.endswith(".md"):
        return raw.decode("utf-8", errors="ignore")

    raise ValueError("仅支持 json/jsonl/csv/xlsx/txt/md 文件")


def chunk_text(text: str, size: int = 500, overlap: int = 80) -> List[str]:
    content = (text or "").strip()
    if not content:
        return []
    chunks = []
    start = 0
    total = len(content)
    while start < total:
        end = min(start + size, total)
        piece = content[start:end].strip()
        if piece:
            chunks.append(piece)
        if end >= total:
            break
        start = max(end - overlap, 0)
    return chunks


def embed_texts(texts: List[str]) -> List[List[float]]:
    api_key = settings.ARK_API_KEY.strip()
    if not api_key:
        raise RuntimeError("ARK_API_KEY 未配置")
    endpoint = (getattr(settings, "ARK_EMBEDDING_ENDPOINT", "/embeddings/multimodal") or "/embeddings/multimodal").strip()
    if not endpoint.startswith("/"):
        endpoint = "/" + endpoint
    url = settings.ARK_EMBEDDING_BASE_URL.rstrip("/") + endpoint
    payload = {"model": settings.ARK_EMBEDDING_MODEL, "input": texts}
    if "multimodal" in endpoint:
        payload = {
            "model": settings.ARK_EMBEDDING_MODEL,
            "instructions": settings.ARK_EMBEDDING_INSTRUCTIONS,
            "input": [{"type": "text", "text": text} for text in texts],
            "dimensions": settings.ARK_EMBEDDING_DIMENSIONS,
            "multi_embedding": {"type": "enabled"},
            "sparse_embedding": {"type": "enabled"},
            "encoding_format": "float",
        }
    timeout_s = max(int(getattr(settings, "ARK_EMBEDDING_TIMEOUT", 8)), 3)
    try:
        resp = requests.post(
            url,
            json=payload,
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            timeout=timeout_s,
        )
    except requests.RequestException as exc:
        raise RuntimeError(f"Embedding 请求失败: {exc}")
    if resp.status_code >= 400:
        raise RuntimeError(f"Embedding 接口错误({resp.status_code})")
    try:
        data = resp.json()
    except ValueError:
        raise RuntimeError("Embedding 返回非JSON格式")
    if not isinstance(data, dict):
        raise RuntimeError(f"Embedding 顶层返回结构异常: {type(data).__name__}")
    items = data.get("data") or []
    if not items:
        raise RuntimeError("Embedding 返回为空")
    if isinstance(items, dict):
        items = [items]
    if not isinstance(items, list):
        raise RuntimeError("Embedding 返回结构异常")

    normalized = []
    for idx, item in enumerate(items):
        if isinstance(item, dict):
            normalized.append(item)
            continue
        # Some providers may return vector list directly.
        if isinstance(item, list) and item and isinstance(item[0], (int, float)):
            normalized.append({"index": idx, "embedding": item})
            continue
        # Skip unsupported item shape, but don't crash on `.get`.
    items = sorted(normalized, key=lambda x: x.get("index", 0))
    if not items:
        raise RuntimeError(f"Embedding 返回结构不支持: {str(data)[:180]}")

    def pick_dense_embedding(item):
        if item.get("embedding"):
            return item.get("embedding")
        if item.get("dense_embedding"):
            return item.get("dense_embedding")
        emb = item.get("embeddings") or {}
        if isinstance(emb, dict):
            if emb.get("dense_embedding"):
                return emb.get("dense_embedding")
            if emb.get("embedding"):
                return emb.get("embedding")
        return None

    vectors = [pick_dense_embedding(item) for item in items]
    vectors = [vec for vec in vectors if vec]
    if len(vectors) != len(texts):
        raise RuntimeError(f"Embedding 返回数量异常: got={len(vectors)} expected={len(texts)}")
    return vectors


def _qdrant_client() -> QdrantClient:
    url = settings.QDRANT_URL.strip()
    if not url:
        raise RuntimeError("QDRANT_URL 未配置")
    key = settings.QDRANT_API_KEY.strip()
    return QdrantClient(url=url, api_key=key or None, timeout=30)


def ensure_collection(vector_size: int):
    client = _qdrant_client()
    collection = settings.QDRANT_COLLECTION
    exists = client.collection_exists(collection)
    if exists:
        return
    client.create_collection(
        collection_name=collection,
        vectors_config=VectorParams(size=vector_size, distance=Distance.COSINE),
    )


def upsert_chunks(document_id: int, chunks: List[str], vectors: List[List[float]], start_index: int = 0) -> List[str]:
    if not chunks:
        return []
    ensure_collection(len(vectors[0]))
    client = _qdrant_client()
    collection = settings.QDRANT_COLLECTION
    points = []
    vector_ids = []
    for idx, (text, vec) in enumerate(zip(chunks, vectors)):
        vector_id = uuid.uuid4().hex
        vector_ids.append(vector_id)
        points.append(
            PointStruct(
                id=vector_id,
                vector=vec,
                payload={"document_id": document_id, "chunk_index": start_index + idx, "text": text},
            )
        )
    client.upsert(collection_name=collection, points=points)
    return vector_ids


def search_context(query: str, top_k: int = 5) -> List[Tuple[str, float]]:
    if not getattr(settings, "AI_CS_ENABLE_RAG", True):
        return []
    query = (query or "").strip()
    if not query:
        return []
    vectors = embed_texts([query])
    query_vec = vectors[0]
    client = _qdrant_client()
    result = client.search(
        collection_name=settings.QDRANT_COLLECTION,
        query_vector=query_vec,
        limit=max(top_k, 1),
        with_payload=True,
    )
    output = []
    for item in result:
        text = (item.payload or {}).get("text") or ""
        score = float(getattr(item, "score", 0.0) or 0.0)
        if text:
            output.append((text, score))
    return output


def has_reliable_context(context_blocks: List[Tuple[str, float]], min_score: Optional[float] = None) -> bool:
    threshold = Decimal(str(min_score if min_score is not None else settings.AI_CS_CONTEXT_MIN_SCORE))
    for _, score in context_blocks or []:
        if Decimal(str(score)) >= threshold:
            return True
    return False


def web_search(query: str) -> List[Dict[str, Any]]:
    search_url = (getattr(settings, "SEARXNG_SEARCH_URL", "") or "").strip()
    basic_user = (getattr(settings, "SEARXNG_BASIC_USER", "") or "").strip()
    basic_pass = (getattr(settings, "SEARXNG_BASIC_PASS", "") or "").strip()
    if not search_url:
        return []
    q = (query or "").strip()
    if not q:
        return []

    last_error = ""
    for _ in range(2):  # first try + one retry
        try:
            resp = requests.get(
                search_url,
                params={"q": q},
                auth=(basic_user, basic_pass),
                timeout=20,
            )
        except requests.RequestException as exc:
            last_error = f"联网检索请求失败: {exc}"
            continue
        if resp.status_code != 200:
            last_error = f"联网检索HTTP错误({resp.status_code})"
            continue
        try:
            body = resp.json()
        except ValueError:
            last_error = "联网检索JSON解析失败"
            continue

        raw_results = []
        if isinstance(body, dict):
            raw_results = body.get("results") or []
            if not raw_results and isinstance(body.get("data"), dict):
                raw_results = body["data"].get("results") or []
        elif isinstance(body, list):
            raw_results = body

        if isinstance(raw_results, dict):
            raw_results = [raw_results]
        if not isinstance(raw_results, list):
            raw_results = []

        output: List[Dict[str, Any]] = []
        for item in raw_results:
            if not isinstance(item, dict):
                continue
            title = str(item.get("title") or "").strip()
            url = str(item.get("url") or item.get("link") or "").strip()
            content = str(item.get("content") or item.get("snippet") or item.get("text") or "").strip()
            score_val = item.get("score", 0.0)
            try:
                score = float(score_val or 0.0)
            except (TypeError, ValueError):
                score = 0.0
            if not content:
                continue
            output.append(
                {
                    "title": title,
                    "url": url,
                    "content": content,
                    "score": score,
                }
            )
        return output
    raise RuntimeError(last_error or "联网检索失败")


def search_web(query: str) -> List[Dict[str, Any]]:
    # Alias for tool registry naming convention.
    return web_search(query)


def stream_llm_answer(messages):
    base_url = settings.AI_CS_LLM_BASE_URL.rstrip("/")
    model = settings.AI_CS_LLM_MODEL
    api_key = settings.AI_CS_LLM_API_KEY.strip()
    if not api_key:
        raise RuntimeError("AI_CS_LLM_API_KEY 未配置")
    url = f"{base_url}/chat/completions"
    headers = {"Content-Type": "application/json"}
    headers["Authorization"] = f"Bearer {api_key}"

    payload = {
        "model": model,
        "messages": messages,
        "stream": True,
        "temperature": 0.5,
    }

    with requests.post(url, json=payload, headers=headers, stream=True, timeout=180) as resp:
        if resp.status_code >= 400:
            raise RuntimeError(f"LLM 服务错误({resp.status_code})")
        for raw in resp.iter_lines(decode_unicode=True):
            if not raw:
                continue
            line = raw.strip()
            if not line.startswith("data:"):
                continue
            data = line[5:].strip()
            if data == "[DONE]":
                break
            try:
                obj = json.loads(data)
            except json.JSONDecodeError:
                continue
            choices = obj.get("choices") or []
            if not choices:
                continue
            delta = choices[0].get("delta") or {}
            content = delta.get("content")
            if content:
                yield content


_IMAGE_INTENT_PATTERNS = [
    r"(生成|做|画|创作).{0,8}(图|图片|海报|插画|封面|头像|壁纸)",
    r"(生图|画图|绘图|出图)",
    r"(image|draw|illustration|poster|cover|wallpaper)",
]


def has_image_generation_intent(text: str) -> bool:
    content = (text or "").strip().lower()
    if not content:
        return False
    return any(re.search(pattern, content, flags=re.IGNORECASE) for pattern in _IMAGE_INTENT_PATTERNS)


def optimize_image_prompt(user_text: str) -> str:
    base = (user_text or "").strip()
    if not base:
        return ""
    api_key = settings.AI_CS_LLM_API_KEY.strip()
    if not api_key:
        return base
    payload = {
        "model": settings.AI_CS_LLM_MODEL,
        "stream": False,
        "temperature": 0.4,
        "messages": [
            {
                "role": "system",
                "content": (
                    "你是专业提示词优化助手。"
                    "把用户生图需求改写成高质量中文提示词，保留核心诉求，补齐画面主体、构图、光线、风格、细节。"
                    "不要解释，不要Markdown，不要多段，只输出最终提示词。"
                ),
            },
            {"role": "user", "content": base},
        ],
    }
    url = f"{settings.AI_CS_LLM_BASE_URL.rstrip('/')}/chat/completions"
    try:
        resp = requests.post(
            url,
            json=payload,
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            timeout=45,
        )
        if resp.status_code >= 400:
            return base
        body = resp.json()
        content = (
            ((body.get("choices") or [{}])[0].get("message") or {}).get("content")
            if isinstance(body, dict)
            else ""
        )
        optimized = (content or "").strip()
        return optimized or base
    except Exception:
        return base


def generate_image_with_ark(prompt: str) -> Dict[str, Any]:
    api_key = settings.ARK_API_KEY.strip()
    if not api_key:
        raise RuntimeError("ARK_API_KEY 未配置")
    text = (prompt or "").strip()
    if not text:
        raise RuntimeError("提示词不能为空")

    url = settings.ARK_IMAGE_BASE_URL.rstrip("/") + "/images/generations"
    def _extract_error_message(resp: requests.Response) -> str:
        try:
            err = resp.json()
        except ValueError:
            text_msg = (resp.text or "").strip()
            return text_msg[:300] if text_msg else f"HTTP {resp.status_code}"
        if isinstance(err, dict):
            if isinstance(err.get("error"), dict):
                detail = err["error"].get("message") or err["error"].get("type")
                if detail:
                    return str(detail)
            for key in ("message", "msg", "detail"):
                if err.get(key):
                    return str(err.get(key))
        return str(err)[:300]

    def _parse_success(resp: requests.Response):
        try:
            data = resp.json()
        except ValueError:
            raise RuntimeError("生图接口返回非JSON格式")
        if not isinstance(data, dict):
            raise RuntimeError("生图接口返回结构异常")
        items = data.get("data") or []
        if not isinstance(items, list) or not items:
            raise RuntimeError("生图返回为空")
        first = items[0] if isinstance(items[0], dict) else {}
        image_url = (first.get("url") or "").strip()
        b64_data = (first.get("b64_json") or "").strip()
        fmt = (settings.ARK_IMAGE_OUTPUT_FORMAT or "png").strip().lower() or "png"
        mime = "image/png" if fmt == "png" else f"image/{fmt}"
        if not image_url and b64_data:
            image_url = f"data:{mime};base64,{b64_data}"
        if not image_url:
            raise RuntimeError("生图结果缺少图片地址")
        return {
            "url": image_url,
            "name": f"ai-image.{fmt}",
            "mime_type": mime,
            "source": "ark",
        }

    raw_fallbacks = getattr(settings, "ARK_IMAGE_FALLBACK_MODELS", "") or ""
    models = [settings.ARK_IMAGE_MODEL]
    models.extend([item.strip() for item in raw_fallbacks.split(",") if item.strip()])
    # Ark docs examples often use this newer model id; keep as final fallback.
    models.append("doubao-seedream-5-0-260128")
    seen = set()
    model_candidates = []
    for model in models:
        if model and model not in seen:
            seen.add(model)
            model_candidates.append(model)

    last_error = ""
    timeout_s = max(int(getattr(settings, "ARK_IMAGE_TIMEOUT", 90)), 20)
    for model in model_candidates:
        payload = {
            "model": model,
            "prompt": text,
            "size": settings.ARK_IMAGE_SIZE,
            "output_format": settings.ARK_IMAGE_OUTPUT_FORMAT,
            "watermark": settings.ARK_IMAGE_WATERMARK,
        }
        try:
            resp = requests.post(
                url,
                json=payload,
                headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                timeout=timeout_s,
            )
        except requests.RequestException as exc:
            last_error = f"请求失败: {exc}"
            continue
        if resp.status_code >= 400:
            last_error = f"模型 {model} 错误({resp.status_code}): {_extract_error_message(resp)}"
            continue
        try:
            return _parse_success(resp)
        except Exception as exc:
            last_error = f"模型 {model} 解析失败: {exc}"
            continue

    raise RuntimeError(last_error or "生图接口调用失败")


def build_system_prompt(
    setting: AICustomerSetting,
    context_blocks: List[Tuple[str, float]],
    web_results: Optional[List[Dict[str, Any]]] = None,
) -> str:
    context_lines = []
    threshold = Decimal(str(settings.AI_CS_CONTEXT_MIN_SCORE))
    for idx, (text, score) in enumerate(context_blocks, start=1):
        if Decimal(str(score)) >= threshold:
            context_lines.append(f"[{idx}] score={score:.4f}\n{text}")
    context_text = "\n\n".join(context_lines) or "无可靠知识库命中"
    web_lines = []
    for idx, item in enumerate(web_results or [], start=1):
        title = (item.get("title") or "").strip()
        url = (item.get("url") or "").strip()
        content = (item.get("content") or "").strip()
        if not content:
            continue
        head = f"[W{idx}]"
        if title:
            head += f" {title}"
        if url:
            head += f" ({url})"
        web_lines.append(f"{head}\n{content}")
    web_text = "\n\n".join(web_lines) or "无联网检索结果"

    return (
        f"{setting.base_prompt}\n"
        f"你的语气风格：{setting.tone_style}\n"
        "请严格遵守：\n"
        "1) 优先依据知识库上下文回答，不得编造事实。\n"
        "2) 当知识库不足时，可结合联网检索结果回答，并尽量引用 [Wn]。\n"
        "3) 若上下文不足，先明确不确定，再给出可执行建议。\n"
        "4) 当仍无法可靠回答时，必须在回复最后追加 [NEED_HUMAN]。\n"
        f"\n知识库上下文:\n{context_text}\n"
        f"\n联网检索结果:\n{web_text}\n"
    )
